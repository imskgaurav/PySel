import openpyxl as xl
import panda as pp

wb = xl.load_workbook("/Users/07.gaurav/Documents/TestExcel.xlsx")
sheet = wb['Sheet1']
cellVal = sheet.cell(1,1)
print(cellVal.value)
print(sheet.max_row)
print(sheet.max_column)

for row in range(1, sheet.max_row+1):
    print(row)
    cell_val = sheet.cell(row,2)
    print(cell_val.value)
print()
    

    
    






 